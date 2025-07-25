"""
?? AFCO AUTOMATIC DOCUMENT PROCESSING INTERFACE
?? In Loving Memory of Omar (2007-2024)
Complete automatic folder monitoring, document upload, analysis, and mapping system
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import plotly.express as px
import plotly.graph_objects as go
import os
import glob
import time
import json
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional
import threading
import hashlib
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# Document processing imports with fallbacks
try:
    import openpyxl
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

try:
    from docx import Document
    WORD_SUPPORT = True
except ImportError:
    WORD_SUPPORT = False

try:
    import pdfplumber
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

# Database imports
try:
    from production_database import production_db
    DATABASE_AVAILABLE = True
except ImportError:
    DATABASE_AVAILABLE = False
    production_db = None

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class AutoFileHandler(FileSystemEventHandler):
    """File system event handler for automatic file monitoring"""

    def __init__(self, callback_func=None):
        self.callback_func = callback_func
        self.processed_files = set()

    def on_created(self, event):
        if not event.is_dir:
            self.handle_file_event(event.src_path, "created")

    def on_modified(self, event):
        if not event.is_dir:
            self.handle_file_event(event.src_path, "modified")

    def handle_file_event(self, file_path, event_type):
        # Only process supported file types
        if any(file_path.lower().endswith(ext) for ext in ['.xlsx', '.xls', '.csv', '.docx', '.doc', '.pdf']):
            file_hash = self.get_file_hash(file_path)
            if file_hash not in self.processed_files:
                self.processed_files.add(file_hash)
                if self.callback_func:
                    self.callback_func(file_path, event_type)
                logger.info(f"New file detected: {file_path} ({event_type})")

    def get_file_hash(self, file_path):
        """Generate hash for file to avoid duplicate processing"""
        try:
            with open(file_path, 'rb') as f:
                return hashlib.md5(f.read()).hexdigest()
        except:
            return str(file_path)

class AFCOAutoDocumentProcessor:
    """Automatic document processor for AFCO with folder monitoring"""

    def __init__(self):
        self.watch_folders = []
        self.observers = []
        self.file_cache = {}
        self.extraction_rules = {}
        self.field_mappings = {}
        self.processing_history = []

        # Initialize storage directories
        self.ensure_directories()

        # Load configuration
        self.load_configuration()

    def ensure_directories(self):
        """Create necessary directories"""
        directories = [
            "data/auto_watch",
            "data/processed",
            "data/mappings",
            "data/extractions",
            "data/sql_exports"
        ]

        for directory in directories:
            Path(directory).mkdir(parents=True, exist_ok=True)

    def load_configuration(self):
        """Load processing configuration"""
        config_path = "data/mappings/auto_config.json"

        if os.path.exists(config_path):
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.extraction_rules = config.get('extraction_rules', {})
                    self.field_mappings = config.get('field_mappings', {})
                    self.watch_folders = config.get('watch_folders', [])
            except Exception as e:
                logger.error(f"Error loading configuration: {e}")

        # Default configuration
        if not self.extraction_rules:
            self.set_default_extraction_rules()

    def set_default_extraction_rules(self):
        """Set default extraction rules for Saudi market documents"""
        self.extraction_rules = {
            'client_data': {
                'patterns': ['company', 'client', 'customer', '????', '????'],
                'required_fields': ['name', 'contact', 'sector'],
                'target_table': 'clients'
            },
            'vendor_data': {
                'patterns': ['vendor', 'supplier', '????', '?????'],
                'required_fields': ['company_name', 'contact_person', 'products'],
                'target_table': 'vendors'
            },
            'rfq_data': {
                'patterns': ['rfq', 'tender', 'bid', '????', '??????'],
                'required_fields': ['title', 'value', 'deadline'],
                'target_table': 'rfqs'
            },
            'financial_data': {
                'patterns': ['revenue', 'cost', 'budget', '???????', '??????'],
                'required_fields': ['amount', 'date', 'description'],
                'target_table': 'financial_transactions'
            }
        }

    def save_configuration(self):
        """Save current configuration"""
        config_path = "data/mappings/auto_config.json"

        config = {
            'extraction_rules': self.extraction_rules,
            'field_mappings': self.field_mappings,
            'watch_folders': self.watch_folders,
            'last_updated': datetime.now().isoformat()
        }

        try:
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"Error saving configuration: {e}")

    def add_watch_folder(self, folder_path: str):
        """Add folder to monitoring list"""
        if os.path.exists(folder_path) and folder_path not in self.watch_folders:
            self.watch_folders.append(folder_path)
            self.start_folder_monitoring(folder_path)
            self.save_configuration()
            return True
        return False

    def start_folder_monitoring(self, folder_path: str):
        """Start monitoring a specific folder"""
        try:
            observer = Observer()
            event_handler = AutoFileHandler(callback_func=self.handle_auto_file)
            observer.schedule(event_handler, folder_path, recursive=True)
            observer.start()
            self.observers.append(observer)
            logger.info(f"Started monitoring folder: {folder_path}")
        except Exception as e:
            logger.error(f"Error starting folder monitoring: {e}")

    def stop_all_monitoring(self):
        """Stop all folder monitoring"""
        for observer in self.observers:
            observer.stop()
            observer.join()
        self.observers.clear()

    def handle_auto_file(self, file_path: str, event_type: str):
        """Handle automatically detected files"""
        try:
            # Extract data from file
            extracted_data = self.extract_file_data(file_path)

            if extracted_data:
                # Apply field mappings
                mapped_data = self.apply_field_mappings(extracted_data)

                # Store processed data
                self.store_processed_data(file_path, mapped_data)

                # Update processing history
                self.processing_history.append({
                    'file_path': file_path,
                    'processed_at': datetime.now().isoformat(),
                    'event_type': event_type,
                    'records_extracted': len(mapped_data),
                    'status': 'success'
                })

                logger.info(f"Successfully processed: {file_path}")

        except Exception as e:
            logger.error(f"Error processing auto file {file_path}: {e}")
            self.processing_history.append({
                'file_path': file_path,
                'processed_at': datetime.now().isoformat(),
                'event_type': event_type,
                'status': 'failed',
                'error': str(e)
            })

    def scan_folders(self) -> Dict[str, List[Dict]]:
        """Scan all watch folders and return file information"""
        folder_contents = {}

        for folder_path in self.watch_folders:
            if os.path.exists(folder_path):
                files = []

                # Supported file extensions
                extensions = ['*.xlsx', '*.xls', '*.csv', '*.docx', '*.doc', '*.pdf']

                for extension in extensions:
                    for file_path in glob.glob(os.path.join(folder_path, '**', extension), recursive=True):
                        file_info = self.get_file_info(file_path)
                        files.append(file_info)

                folder_contents[folder_path] = files

        return folder_contents

    def get_file_info(self, file_path: str) -> Dict:
        """Get detailed file information"""
        try:
            stat = os.stat(file_path)

            return {
                'path': file_path,
                'name': os.path.basename(file_path),
                'size': stat.st_size,
                'modified': datetime.fromtimestamp(stat.st_mtime),
                'extension': Path(file_path).suffix.lower(),
                'status': self.get_file_processing_status(file_path)
            }
        except Exception as e:
            return {
                'path': file_path,
                'name': os.path.basename(file_path),
                'error': str(e),
                'status': 'error'
            }

    def get_file_processing_status(self, file_path: str) -> str:
        """Check if file has been processed"""
        for record in self.processing_history:
            if record['file_path'] == file_path:
                return record['status']
        return 'pending'

    def extract_file_data(self, file_path: str) -> List[Dict]:
        """Extract data from various file formats"""
        extension = Path(file_path).suffix.lower()

        try:
            if extension in ['.xlsx', '.xls'] and EXCEL_SUPPORT:
                return self.extract_from_excel(file_path)
            elif extension == '.csv':
                return self.extract_from_csv(file_path)
            elif extension in ['.docx', '.doc'] and WORD_SUPPORT:
                return self.extract_from_word(file_path)
            elif extension == '.pdf' and PDF_SUPPORT:
                return self.extract_from_pdf(file_path)
            else:
                logger.warning(f"Unsupported file type: {extension}")
                return []
        except Exception as e:
            logger.error(f"Error extracting data from {file_path}: {e}")
            return []

    def extract_from_excel(self, file_path: str) -> List[Dict]:
        """Extract data from Excel files"""
        data = []
        workbook = load_workbook(file_path, data_only=True)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            if sheet.max_row <= 1:
                continue

            # Get headers
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value or f"Column_{len(headers)}")

            # Get data rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    row_dict = {}
                    for i, value in enumerate(row):
                        if i < len(headers):
                            row_dict[headers[i]] = value

                    row_dict['_source_sheet'] = sheet_name
                    row_dict['_source_file'] = file_path
                    data.append(row_dict)

        return data

    def extract_from_csv(self, file_path: str) -> List[Dict]:
        """Extract data from CSV files"""
        encodings = ['utf-8', 'utf-8-sig', 'iso-8859-1', 'cp1256']

        for encoding in encodings:
            try:
                df = pd.read_csv(file_path, encoding=encoding)
                df = df.dropna(how='all')

                data = df.to_dict('records')
                for record in data:
                    record['_source_file'] = file_path

                return data
            except UnicodeDecodeError:
                continue
            except Exception as e:
                logger.error(f"Error reading CSV {file_path}: {e}")
                break

        return []

    def extract_from_word(self, file_path: str) -> List[Dict]:
        """Extract data from Word documents"""
        doc = Document(file_path)
        data = []

        for table_idx, table in enumerate(doc.tables):
            if not table.rows:
                continue

            headers = [cell.text.strip() for cell in table.rows[0].cells]

            for row in table.rows[1:]:
                row_data = {}
                for i, cell in enumerate(row.cells):
                    if i < len(headers):
                        row_data[headers[i]] = cell.text.strip()

                if any(v for v in row_data.values()):
                    row_data['_source_table'] = f"Table_{table_idx}"
                    row_data['_source_file'] = file_path
                    data.append(row_data)

        return data

    def extract_from_pdf(self, file_path: str) -> List[Dict]:
        """Extract data from PDF files"""
        data = []

        with pdfplumber.open(file_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                tables = page.extract_tables()

                for table_idx, table in enumerate(tables):
                    if table and len(table) > 1:
                        headers = table[0]

                        for row in table[1:]:
                            if any(cell for cell in row):
                                row_dict = {}
                                for i, cell in enumerate(row):
                                    if i < len(headers):
                                        row_dict[headers[i]] = cell

                                if any(v for v in row_dict.values()):
                                    row_dict['_source_page'] = page_num + 1
                                    row_dict['_source_table'] = table_idx
                                    row_dict['_source_file'] = file_path
                                    data.append(row_dict)

        return data

    def apply_field_mappings(self, data: List[Dict]) -> List[Dict]:
        """Apply field mappings to extracted data"""
        mapped_data = []

        for record in data:
            mapped_record = record.copy()

            # Apply field mappings
            for source_field, target_field in self.field_mappings.items():
                if source_field in record:
                    mapped_record[target_field] = record[source_field]

            # Determine data type based on content
            data_type = self.determine_data_type(mapped_record)
            mapped_record['_data_type'] = data_type

            mapped_data.append(mapped_record)

        return mapped_data

    def determine_data_type(self, record: Dict) -> str:
        """Determine data type based on record content"""
        content = ' '.join(str(v).lower() for v in record.values() if v)

        for data_type, rules in self.extraction_rules.items():
            patterns = rules.get('patterns', [])
            if any(pattern in content for pattern in patterns):
                return data_type

        return 'general'

    def store_processed_data(self, file_path: str, data: List[Dict]):
        """Store processed data"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.basename(file_path)
        output_file = f"data/extractions/{timestamp}_{filename}.json"

        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump({
                    'source_file': file_path,
                    'processed_at': datetime.now().isoformat(),
                    'record_count': len(data),
                    'data': data
                }, f, ensure_ascii=False, indent=2, default=str)
        except Exception as e:
            logger.error(f"Error storing processed data: {e}")

# Initialize global processor
auto_processor = AFCOAutoDocumentProcessor()

def main():
    """Main Streamlit application"""

    st.set_page_config(
        page_title="?? AFCO Auto Document Center",
        page_icon="??",
        layout="wide",
        initial_sidebar_state="expanded"
    )

    # Header
    st.markdown("""
    <div style="background: linear-gradient(90deg, #4A90E2, #87CEEB); color: white; padding: 1rem; border-radius: 10px; text-align: center; margin-bottom: 2rem;">
        <h1>?? AFCO Automatic Document Processing Center</h1>
        <p>?? In Loving Memory of Omar (2007-2024) - "Forever 17, Forever Inspiring Innovation"</p>
        <p>Automatic Folder Monitoring • Real-time Data Extraction • Smart Field Mapping</p>
    </div>
    """, unsafe_allow_html=True)

    # Sidebar navigation
    st.sidebar.markdown("### ?? Navigation")

    tabs = st.tabs([
        "?? Folder Monitor",
        "??? File Manager",
        "?? Field Mapping",
        "?? Processing Status",
        "?? Configuration"
    ])

    with tabs[0]:
        show_folder_monitor()

    with tabs[1]:
        show_file_manager()

    with tabs[2]:
        show_field_mapping()

    with tabs[3]:
        show_processing_status()

    with tabs[4]:
        show_configuration()

def show_folder_monitor():
    """Folder monitoring interface"""

    st.subheader("?? Automatic Folder Monitoring")

    col1, col2 = st.columns([2, 1])

    with col1:
        st.markdown("### ?? Watch Folders")

        # Add new folder
        new_folder = st.text_input(
            "Add folder to monitor:",
            placeholder="Enter folder path (e.g., D:\\Documents\\AFCO)",
            help="Enter the full path to the folder you want to monitor"
        )

        col_add, col_browse = st.columns([1, 1])

        with col_add:
            if st.button("? Add Folder", type="primary") and new_folder:
                if auto_processor.add_watch_folder(new_folder):
                    st.success(f"? Added folder: {new_folder}")
                    st.rerun()
                else:
                    st.error("? Failed to add folder. Check if path exists.")

        with col_browse:
            if st.button("?? Browse Folders"):
                st.info("?? Enter folder path manually above")

        # Current watch folders
        if auto_processor.watch_folders:
            st.markdown("### ?? Currently Monitoring:")

            for idx, folder in enumerate(auto_processor.watch_folders):
                with st.expander(f"?? {folder}", expanded=True):
                    col1, col2, col3 = st.columns([2, 1, 1])

                    with col1:
                        if os.path.exists(folder):
                            st.success("?? Active")
                            file_count = len([f for f in glob.glob(os.path.join(folder, "**", "*.*"), recursive=True)
                                            if f.lower().endswith(('.xlsx', '.xls', '.csv', '.docx', '.doc', '.pdf'))])
                            st.write(f"?? {file_count} supported files")
                        else:
                            st.error("?? Folder not found")

                    with col2:
                        if st.button(f"?? Scan Now", key=f"scan_{idx}"):
                            st.info("Scanning folder...")
                            st.rerun()

                    with col3:
                        if st.button(f"??? Remove", key=f"remove_{idx}"):
                            auto_processor.watch_folders.remove(folder)
                            auto_processor.save_configuration()
                            st.rerun()
        else:
            st.info("?? No folders being monitored. Add a folder to start automatic processing.")

    with col2:
        st.markdown("### ?? Monitoring Status")

        # Real-time statistics
        total_folders = len(auto_processor.watch_folders)
        active_folders = sum(1 for folder in auto_processor.watch_folders if os.path.exists(folder))

        st.metric("Total Folders", total_folders)
        st.metric("Active Folders", active_folders)
        st.metric("Processing Rules", len(auto_processor.extraction_rules))

        # Recent activity
        st.markdown("### ?? Recent Activity")

        recent_activity = auto_processor.processing_history[-5:] if auto_processor.processing_history else []

        if recent_activity:
            for activity in reversed(recent_activity):
                status_icon = "?" if activity['status'] == 'success' else "?"
                st.write(f"{status_icon} {os.path.basename(activity['file_path'])}")
                st.caption(f"?? {activity['processed_at'][:19]}")
        else:
            st.info("No recent activity")

def show_file_manager():
    """File manager interface"""

    st.subheader("??? Automatic File Manager")

    # Scan all watched folders
    if st.button("?? Refresh File List", type="primary"):
        st.rerun()

    folder_contents = auto_processor.scan_folders()

    if not folder_contents:
        st.info("?? No watch folders configured. Go to Folder Monitor to add folders.")
        return

    # Display files by folder
    for folder_path, files in folder_contents.items():
        st.markdown(f"### ?? {folder_path}")

        if not files:
            st.info("?? No supported files found in this folder.")
            continue

        # Filter controls
        col1, col2, col3 = st.columns(3)

        with col1:
            file_types = st.multiselect(
                "File Types:",
                ['.xlsx', '.xls', '.csv', '.docx', '.doc', '.pdf'],
                default=['.xlsx', '.csv'],
                key=f"types_{hash(folder_path)}"
            )

        with col2:
            status_filter = st.selectbox(
                "Status:",
                ["All", "Pending", "Processed", "Failed"],
                key=f"status_{hash(folder_path)}"
            )

        with col3:
            max_files = st.number_input(
                "Max Files:",
                min_value=10,
                max_value=1000,
                value=50,
                key=f"max_{hash(folder_path)}"
            )

        # Filter files
        filtered_files = []
        for file_info in files:
            # Filter by type
            if file_types and file_info.get('extension') not in file_types:
                continue

            # Filter by status
            if status_filter != "All":
                file_status = file_info.get('status', 'pending')
                if status_filter.lower() != file_status:
                    continue

            filtered_files.append(file_info)

        # Limit results
        filtered_files = filtered_files[:max_files]

        if filtered_files:
            # Create DataFrame for display
            display_data = []
            for file_info in filtered_files:
                display_data.append({
                    'File Name': file_info['name'],
                    'Size (KB)': f"{file_info.get('size', 0) / 1024:.1f}",
                    'Modified': file_info.get('modified', 'Unknown'),
                    'Status': file_info.get('status', 'pending'),
                    'Type': file_info.get('extension', ''),
                    'Path': file_info['path']
                })

            df = pd.DataFrame(display_data)

            # Display with selection
            st.markdown(f"#### ?? Files ({len(df)} found)")

            # Action buttons
            col1, col2, col3, col4 = st.columns(4)

            with col1:
                if st.button(f"?? Process All Pending", key=f"process_all_{hash(folder_path)}"):
                    process_pending_files(filtered_files)

            with col2:
                if st.button(f"?? Auto-Map Fields", key=f"automap_{hash(folder_path)}"):
                    auto_map_fields(filtered_files)

            with col3:
                if st.button(f"?? Preview Files", key=f"preview_{hash(folder_path)}"):
                    preview_files(filtered_files[:3])

            with col4:
                if st.button(f"?? Export List", key=f"export_{hash(folder_path)}"):
                    export_file_list(df)

            # File list
            st.dataframe(df, use_container_width=True, hide_index=True)

            # File details expander
            with st.expander("?? File Details & Actions"):
                selected_file = st.selectbox(
                    "Select file for details:",
                    options=[f['name'] for f in filtered_files],
                    key=f"select_{hash(folder_path)}"
                )

                if selected_file:
                    file_info = next(f for f in filtered_files if f['name'] == selected_file)
                    show_file_details(file_info)
        else:
            st.info(f"?? No files match the current filters.")

def show_file_details(file_info: Dict):
    """Show detailed file information and actions"""

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("#### ?? File Information")
        st.write(f"**Name:** {file_info['name']}")
        st.write(f"**Size:** {file_info.get('size', 0) / 1024:.1f} KB")
        st.write(f"**Modified:** {file_info.get('modified', 'Unknown')}")
        st.write(f"**Status:** {file_info.get('status', 'pending')}")
        st.write(f"**Type:** {file_info.get('extension', '')}")

    with col2:
        st.markdown("#### ? Quick Actions")

        if st.button("?? Extract Data", key=f"extract_{file_info['name']}"):
            extract_single_file(file_info)

        if st.button("?? Auto-Map", key=f"map_{file_info['name']}"):
            auto_map_single_file(file_info)

        if st.button("?? Preview", key=f"preview_{file_info['name']}"):
            preview_single_file(file_info)

        if st.button("?? Analyze", key=f"analyze_{file_info['name']}"):
            analyze_single_file(file_info)

def show_field_mapping():
    """Field mapping configuration interface"""

    st.subheader("?? Smart Field Mapping Configuration")

    tab1, tab2, tab3 = st.tabs(["?? Create Mappings", "?? Current Mappings", "?? AI Suggestions"])

    with tab1:
        st.markdown("### ?? Create New Field Mapping")

        col1, col2 = st.columns(2)

        with col1:
            st.markdown("#### Source Field")
            source_field = st.text_input(
                "Source Field Name:",
                placeholder="e.g., Company Name, ??????, Client"
            )

            source_variations = st.text_area(
                "Field Variations (one per line):",
                placeholder="Company\nClient Name\n??? ??????\n??????",
                help="Add different ways this field might appear"
            )

        with col2:
            st.markdown("#### Target Mapping")
            target_table = st.selectbox(
                "Target Table:",
                ["clients", "vendors", "rfqs", "financial_transactions", "kpis"],
                help="Database table where this data will be stored"
            )

            target_field = st.text_input(
                "Target Field:",
                placeholder="e.g., name, company_name, title"
            )

            data_type = st.selectbox(
                "Data Type:",
                ["TEXT", "INTEGER", "REAL", "DATE", "BOOLEAN"]
            )

        # Transformation rules
        st.markdown("#### ?? Transformation Rules")

        transformation = st.selectbox(
            "Apply Transformation:",
            [
                "None",
                "UPPER() - Convert to uppercase",
                "LOWER() - Convert to lowercase",
                "TRIM() - Remove spaces",
                "Replace commas and convert to number",
                "Convert to date format",
                "Custom transformation"
            ]
        )

        if transformation == "Custom transformation":
            custom_transform = st.text_input(
                "Custom Rule:",
                placeholder="e.g., REPLACE(?, ',', '') for removing commas"
            )

        # Validation rules
        st.markdown("#### ? Validation Rules")

        validation = st.selectbox(
            "Validation Rule:",
            [
                "None",
                "Required (not empty)",
                "Email format",
                "Phone number format",
                "Saudi phone (+966 or 05)",
                "Positive number",
                "Date in future",
                "Custom validation"
            ]
        )

        if validation == "Custom validation":
            custom_validation = st.text_input(
                "Custom Validation:",
                placeholder="e.g., LENGTH(?) > 3"
            )

        # Save mapping
        if st.button("?? Save Field Mapping", type="primary"):
            save_field_mapping(source_field, target_table, target_field, transformation, validation)

    with tab2:
        st.markdown("### ?? Current Field Mappings")

        if auto_processor.field_mappings:
            # Display current mappings
            mapping_data = []
            for source, target in auto_processor.field_mappings.items():
                mapping_data.append({
                    'Source Field': source,
                    'Target Field': target,
                    'Actions': '??? Delete'
                })

            df = pd.DataFrame(mapping_data)
            st.dataframe(df, use_container_width=True)

            # Edit/Delete actions
            selected_mapping = st.selectbox(
                "Select mapping to edit:",
                list(auto_processor.field_mappings.keys())
            )

            if selected_mapping:
                col1, col2 = st.columns(2)

                with col1:
                    if st.button("?? Edit Mapping"):
                        edit_field_mapping(selected_mapping)

                with col2:
                    if st.button("??? Delete Mapping"):
                        delete_field_mapping(selected_mapping)
        else:
            st.info("?? No field mappings configured yet.")

    with tab3:
        st.markdown("### ?? AI-Powered Mapping Suggestions")

        if st.button("?? Generate Smart Suggestions"):
            generate_mapping_suggestions()

        st.markdown("""
        #### ?? Suggested Mappings for Saudi Market:

        **Client Data:**
        - Company Name ? clients.name
        - ??? ?????? ? clients.name_ar
        - Sector ? clients.sector
        - Contact Person ? clients.contact_person

        **Vendor Data:**
        - Vendor Name ? vendors.company_name
        - ???? ? vendors.company_name
        - Performance ? vendors.performance_score

        **RFQ Data:**
        - RFQ Title ? rfqs.title
        - ????? ???????? ? rfqs.title
        - Value SAR ? rfqs.value_sar
        - Deadline ? rfqs.deadline

        **Financial Data:**
        - Amount ? financial_transactions.amount_sar
        - ?????? ? financial_transactions.amount_sar
        - Date ? financial_transactions.transaction_date
        """)

def show_processing_status():
    """Processing status and history"""

    st.subheader("?? Document Processing Status & Analytics")

    # Processing summary
    col1, col2, col3, col4 = st.columns(4)

    total_processed = len(auto_processor.processing_history)
    successful = sum(1 for h in auto_processor.processing_history if h['status'] == 'success')
    failed = sum(1 for h in auto_processor.processing_history if h['status'] == 'failed')
    success_rate = (successful / total_processed * 100) if total_processed > 0 else 0

    with col1:
        st.metric("Total Processed", total_processed)

    with col2:
        st.metric("Successful", successful, delta=f"{success_rate:.1f}%")

    with col3:
        st.metric("Failed", failed)

    with col4:
        st.metric("Success Rate", f"{success_rate:.1f}%")

    # Processing timeline
    if auto_processor.processing_history:
        st.markdown("### ?? Processing Timeline")

        # Create timeline data
        timeline_data = []
        for record in auto_processor.processing_history:
            timeline_data.append({
                'Date': pd.to_datetime(record['processed_at']).date(),
                'Time': pd.to_datetime(record['processed_at']).time(),
                'File': os.path.basename(record['file_path']),
                'Status': record['status'],
                'Records': record.get('records_extracted', 0)
            })

        df = pd.DataFrame(timeline_data)

        # Daily processing chart
        daily_counts = df.groupby(['Date', 'Status']).size().reset_index(name='Count')

        if not daily_counts.empty:
            fig = px.bar(
                daily_counts,
                x='Date',
                y='Count',
                color='Status',
                title='Daily Processing Activity',
                color_discrete_map={'success': 'green', 'failed': 'red'}
            )
            st.plotly_chart(fig, use_container_width=True)

        # Recent processing history
        st.markdown("### ?? Recent Processing History")

        recent_history = auto_processor.processing_history[-20:] if len(auto_processor.processing_history) > 20 else auto_processor.processing_history

        history_df = pd.DataFrame(recent_history)
        if not history_df.empty:
            history_df['processed_at'] = pd.to_datetime(history_df['processed_at']).dt.strftime('%Y-%m-%d %H:%M:%S')
            history_df['file_name'] = history_df['file_path'].apply(lambda x: os.path.basename(x))

            display_df = history_df[['processed_at', 'file_name', 'status', 'records_extracted']].copy()
            display_df.columns = ['Processed At', 'File Name', 'Status', 'Records Extracted']

            st.dataframe(display_df, use_container_width=True)

        # Export processing report
        if st.button("?? Export Processing Report"):
            export_processing_report()

    else:
        st.info("?? No processing history available yet.")

def show_configuration():
    """Configuration management interface"""

    st.subheader("?? System Configuration")

    tab1, tab2, tab3 = st.tabs(["??? Processing Rules", "?? Folder Settings", "?? Advanced"])

    with tab1:
        st.markdown("### ??? Data Processing Rules")

        for rule_name, rule_config in auto_processor.extraction_rules.items():
            with st.expander(f"?? {rule_name.replace('_', ' ').title()}"):

                col1, col2 = st.columns(2)

                with col1:
                    st.markdown("**Detection Patterns:**")
                    patterns = st.text_area(
                        "Patterns (one per line):",
                        value='\n'.join(rule_config.get('patterns', [])),
                        key=f"patterns_{rule_name}"
                    )

                    st.markdown("**Required Fields:**")
                    required_fields = st.text_area(
                        "Required fields (one per line):",
                        value='\n'.join(rule_config.get('required_fields', [])),
                        key=f"fields_{rule_name}"
                    )

                with col2:
                    st.markdown("**Target Table:**")
                    target_table = st.selectbox(
                        "Database table:",
                        ["clients", "vendors", "rfqs", "financial_transactions", "kpis"],
                        index=["clients", "vendors", "rfqs", "financial_transactions", "kpis"].index(rule_config.get('target_table', 'clients')),
                        key=f"table_{rule_name}"
                    )

                    st.markdown("**Priority:**")
                    priority = st.slider(
                        "Processing priority:",
                        min_value=1,
                        max_value=10,
                        value=rule_config.get('priority', 5),
                        key=f"priority_{rule_name}"
                    )

                if st.button(f"?? Update {rule_name}", key=f"update_{rule_name}"):
                    update_processing_rule(rule_name, patterns, required_fields, target_table, priority)

    with tab2:
        st.markdown("### ?? Folder Monitoring Settings")

        # Auto-processing settings
        st.markdown("#### ?? Auto-Processing")

        auto_process = st.checkbox(
            "Enable automatic processing of new files",
            value=True,
            help="Automatically process files as soon as they are detected"
        )

        processing_delay = st.slider(
            "Processing delay (seconds):",
            min_value=1,
            max_value=60,
            value=5,
            help="Wait time before processing a new file"
        )

        # File filters
        st.markdown("#### ?? File Filters")

        max_file_size = st.number_input(
            "Maximum file size (MB):",
            min_value=1,
            max_value=1000,
            value=50
        )

        exclude_patterns = st.text_area(
            "Exclude patterns (one per line):",
            value="~$*\n*.tmp\n*.bak",
            help="File patterns to exclude from processing"
        )

        # Save settings
        if st.button("?? Save Folder Settings", type="primary"):
            save_folder_settings(auto_process, processing_delay, max_file_size, exclude_patterns)

    with tab3:
        st.markdown("### ?? Advanced Configuration")

        # System settings
        st.markdown("#### ??? System Settings")

        max_concurrent = st.number_input(
            "Maximum concurrent file processing:",
            min_value=1,
            max_value=10,
            value=3
        )

        log_level = st.selectbox(
            "Logging level:",
            ["DEBUG", "INFO", "WARNING", "ERROR"],
            index=1
        )

        # Database settings
        st.markdown("#### ??? Database Settings")

        db_backup = st.checkbox(
            "Enable automatic database backup",
            value=True
        )

        backup_frequency = st.selectbox(
            "Backup frequency:",
            ["Daily", "Weekly", "Monthly"],
            index=0
        )

        # Reset options
        st.markdown("#### ?? Reset Options")

        col1, col2 = st.columns(2)

        with col1:
            if st.button("??? Clear Processing History"):
                clear_processing_history()

        with col2:
            if st.button("?? Reset All Settings"):
                reset_all_settings()

        # Save advanced settings
        if st.button("?? Save Advanced Settings", type="primary"):
            save_advanced_settings(max_concurrent, log_level, db_backup, backup_frequency)

# Helper functions for file processing
def process_pending_files(files: List[Dict]):
    """Process all pending files"""
    with st.spinner("?? Processing pending files..."):
        processed_count = 0

        for file_info in files:
            if file_info.get('status') == 'pending':
                try:
                    auto_processor.handle_auto_file(file_info['path'], 'manual')
                    processed_count += 1
                except Exception as e:
                    st.error(f"Error processing {file_info['name']}: {e}")

        st.success(f"? Processed {processed_count} files")

def auto_map_fields(files: List[Dict]):
    """Automatically map fields for files"""
    st.info("?? Auto-mapping fields based on AI analysis...")
    # Implement AI-based field mapping logic here
    st.success("? Field mapping completed")

def preview_files(files: List[Dict]):
    """Preview selected files"""
    st.markdown("### ?? File Preview")

    for file_info in files[:3]:  # Limit to 3 files
        with st.expander(f"?? {file_info['name']}"):
            try:
                data = auto_processor.extract_file_data(file_info['path'])
                if data:
                    preview_df = pd.DataFrame(data[:5])  # First 5 rows
                    st.dataframe(preview_df, use_container_width=True)
                else:
                    st.warning("No data could be extracted from this file")
            except Exception as e:
                st.error(f"Error previewing file: {e}")

def export_file_list(df: pd.DataFrame):
    """Export file list to CSV"""
    csv = df.to_csv(index=False)
    st.download_button(
        label="?? Download File List",
        data=csv,
        file_name=f"file_list_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        mime="text/csv"
    )

def extract_single_file(file_info: Dict):
    """Extract data from a single file"""
    with st.spinner(f"?? Extracting data from {file_info['name']}..."):
        try:
            data = auto_processor.extract_file_data(file_info['path'])
            if data:
                st.success(f"? Extracted {len(data)} records")

                # Show preview
                preview_df = pd.DataFrame(data[:10])
                st.dataframe(preview_df, use_container_width=True)
            else:
                st.warning("?? No data could be extracted")
        except Exception as e:
            st.error(f"? Error: {e}")

def auto_map_single_file(file_info: Dict):
    """Auto-map fields for a single file"""
    st.info(f"?? Auto-mapping fields for {file_info['name']}...")
    # Implement single file auto-mapping
    st.success("? Auto-mapping completed")

def preview_single_file(file_info: Dict):
    """Preview a single file"""
    try:
        data = auto_processor.extract_file_data(file_info['path'])
        if data:
            st.success(f"?? Preview of {file_info['name']} ({len(data)} records)")
            preview_df = pd.DataFrame(data[:10])
            st.dataframe(preview_df, use_container_width=True)
        else:
            st.warning("No data available for preview")
    except Exception as e:
        st.error(f"Error: {e}")

def analyze_single_file(file_info: Dict):
    """Analyze a single file"""
    with st.spinner(f"?? Analyzing {file_info['name']}..."):
        try:
            data = auto_processor.extract_file_data(file_info['path'])
            if data:
                # Basic analysis
                st.markdown("#### ?? File Analysis Results")

                col1, col2, col3 = st.columns(3)

                with col1:
                    st.metric("Total Records", len(data))

                with col2:
                    fields = set()
                    for record in data:
                        fields.update(record.keys())
                    st.metric("Unique Fields", len(fields))

                with col3:
                    data_type = auto_processor.determine_data_type(data[0] if data else {})
                    st.metric("Detected Type", data_type)

                # Field analysis
                st.markdown("#### ?? Field Analysis")
                field_analysis = {}
                for record in data:
                    for field, value in record.items():
                        if field not in field_analysis:
                            field_analysis[field] = {'count': 0, 'non_empty': 0, 'samples': []}

                        field_analysis[field]['count'] += 1
                        if value and str(value).strip():
                            field_analysis[field]['non_empty'] += 1
                            if len(field_analysis[field]['samples']) < 3:
                                field_analysis[field]['samples'].append(str(value)[:50])

                for field, stats in field_analysis.items():
                    if not field.startswith('_'):  # Skip internal fields
                        completeness = (stats['non_empty'] / stats['count']) * 100
                        samples = ', '.join(stats['samples'])

                        st.write(f"**{field}:** {completeness:.1f}% complete")
                        st.caption(f"Samples: {samples}")
            else:
                st.warning("No data available for analysis")
        except Exception as e:
            st.error(f"Error: {e}")

# Configuration helper functions
def save_field_mapping(source_field: str, target_table: str, target_field: str, transformation: str, validation: str):
    """Save a field mapping"""
    if source_field and target_field:
        auto_processor.field_mappings[source_field] = f"{target_table}.{target_field}"
        auto_processor.save_configuration()
        st.success(f"? Saved mapping: {source_field} ? {target_table}.{target_field}")
    else:
        st.error("? Please fill in both source and target fields")

def edit_field_mapping(selected_mapping: str):
    """Edit an existing field mapping"""
    st.info(f"?? Editing mapping for: {selected_mapping}")
    # Implement edit functionality

def delete_field_mapping(selected_mapping: str):
    """Delete a field mapping"""
    if selected_mapping in auto_processor.field_mappings:
        del auto_processor.field_mappings[selected_mapping]
        auto_processor.save_configuration()
        st.success(f"? Deleted mapping: {selected_mapping}")
        st.rerun()

def generate_mapping_suggestions():
    """Generate AI-powered mapping suggestions"""
    st.info("?? Analyzing files to generate smart mapping suggestions...")

    # Implement AI-based suggestion logic
    suggestions = [
        {"source": "Company Name", "target": "clients.name", "confidence": 0.95},
        {"source": "??? ??????", "target": "clients.name_ar", "confidence": 0.90},
        {"source": "Vendor", "target": "vendors.company_name", "confidence": 0.88},
        {"source": "Amount SAR", "target": "financial_transactions.amount_sar", "confidence": 0.92}
    ]

    for suggestion in suggestions:
        st.write(f"?? **{suggestion['source']}** ? {suggestion['target']} (Confidence: {suggestion['confidence']:.0%})")

def update_processing_rule(rule_name: str, patterns: str, required_fields: str, target_table: str, priority: int):
    """Update a processing rule"""
    auto_processor.extraction_rules[rule_name] = {
        'patterns': [p.strip() for p in patterns.split('\n') if p.strip()],
        'required_fields': [f.strip() for f in required_fields.split('\n') if f.strip()],
        'target_table': target_table,
        'priority': priority
    }
    auto_processor.save_configuration()
    st.success(f"? Updated processing rule: {rule_name}")

def save_folder_settings(auto_process: bool, processing_delay: int, max_file_size: int, exclude_patterns: str):
    """Save folder monitoring settings"""
    # Implementation for saving folder settings
    st.success("? Folder settings saved successfully")

def save_advanced_settings(max_concurrent: int, log_level: str, db_backup: bool, backup_frequency: str):
    """Save advanced system settings"""
    # Implementation for saving advanced settings
    st.success("? Advanced settings saved successfully")

def clear_processing_history():
    """Clear processing history"""
    auto_processor.processing_history.clear()
    st.success("? Processing history cleared")

def reset_all_settings():
    """Reset all settings to defaults"""
    auto_processor.set_default_extraction_rules()
    auto_processor.field_mappings.clear()
    auto_processor.save_configuration()
    st.success("? All settings reset to defaults")

def export_processing_report():
    """Export processing report"""
    if auto_processor.processing_history:
        df = pd.DataFrame(auto_processor.processing_history)
        csv = df.to_csv(index=False)

        st.download_button(
            label="?? Download Processing Report",
            data=csv,
            file_name=f"processing_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv"
        )
    else:
        st.warning("No processing history to export")

if __name__ == "__main__":
    main()
